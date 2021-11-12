from openpyxl import *


class Eexcel:
    def __init__(self):
        pass

    def save(self, a, b, c):
        wb = load_workbook('aexcel.xlsx')
        ws = wb.active
        ws.append([a,b,c])
        wb.save('aexcel.xlsx')
        wb.close()

    def load(self):
        wb = load_workbook('aexcel.xlsx')
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                print(cell.value, end=" ")
            print()
        wb.save('aexcel.xlsx')
        wb.close()

    def create(self):
        print("새로 만들기")
        wb = Workbook()
        wb.save('aexcel.xlsx')
        wb.close()