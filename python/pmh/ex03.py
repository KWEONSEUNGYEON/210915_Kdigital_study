# def doA(a,b,c, d=20):
#     print("a = ", a, "b = ", b, "c = ", c)
#     print("d = ", d)
# doA(c=20, b=10, a=5, b)

from random import randint
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title='제목제목'

ws['A1'] = 10
ws['A2'] = 20
ws['A3'] = 30

ws['b1'] = 40
ws['b2'] = 50
ws['b3'] = 60

print(ws['A2'])
print(ws['A2'].value)
print(ws['B3'].value)

ws.cell(column=1, row=1).value =100
ws.cell(column=2, row=1).value =200

ws.cell(2,1,300)

index = 1
for x in range(1,11):
    for y in range(1,11):
        ws.cell(x,y,randint(1,100))
        index+=1

wb.save("c.xlsx")
wb.close()